# -*- coding: utf-8 -*-
'''
@File    :   get_caseinfo.py    
@Contact :   272575173@qq.com
@Author  :   Wuyz  
@Time    :   2020/5/8 13:49  
'''

from openpyxl import load_workbook

#打开excel
wb_test=load_workbook(r'C:\Users\Administrator\Desktop\test.xlsx')
#定位sheet
ws_test=wb_test['case']
#获取单元格的值
value=ws_test.cell(1,1).value
#获取最大行列
max_rows=ws_test.max_row
max_col=ws_test.max_column
print(max_col,max_rows)

#获取第一行的值 放入列表字典
test_data=[]
for row in range(1,max_rows+1):  #(1,6) 1 2 3 4 5 6
    sub_data={}
    sub_data['case_name']=ws_test.cell(row+1,1).value
    sub_data['method']=ws_test.cell(row+1,2).value
    sub_data['url']=ws_test.cell(row+1,3).value
    sub_data['data']=ws_test.cell(row+1,4).value
    sub_data['headers']=ws_test.cell(row+1,5).value
    sub_data['assert']=ws_test.cell(row+1,6).value
    test_data.append(sub_data)
    print(sub_data)
print(test_data)

#获取列表所有数据 放入列表字典
# test_data=[]
# for row in range(2,max_rows+1):
#     dict_data = {}
#     for col in range(1,max_col+1):
#         value=ws_test.cell(row,col).value
#         dict_data[ws_test.cell(1,col).value]=value
#     test_data.append(dict_data)
#     print(dict_data)
# print(test_data)
#




